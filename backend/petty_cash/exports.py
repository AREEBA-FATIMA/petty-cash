from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from decimal import Decimal


HEADER_FILL = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='ffffff', size=11)
SUBHEADER_FILL = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=10)
BODY_FONT = Font(name='Calibri', size=10)
MONEY_FORMAT = '#,##0.00'
DATE_FORMAT = 'YYYY-MM-DD'
THIN_BORDER = Border(
    left=Side(style='thin', color='e2e8f0'),
    right=Side(style='thin', color='e2e8f0'),
    top=Side(style='thin', color='e2e8f0'),
    bottom=Side(style='thin', color='e2e8f0'),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_row(ws, row, cols, money_cols=None):
    money_cols = money_cols or []
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal='right' if col in money_cols else 'left',
            vertical='center'
        )
        if col in money_cols and cell.value:
            cell.number_format = MONEY_FORMAT


def auto_width(ws, cols):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)


def export_cashbook(transactions, replenishments, year, month):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Cashbook {month:02d}-{year}'

    headers = ['Date', 'Particulars', 'Category / Ref', 'Branch', 'Vendor',
               'Debit (Out)', 'Credit (In)', 'Receipt', 'Status']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    # Merge all entries
    entries = []
    for t in transactions:
        entries.append({
            'date': t.date.isoformat(),
            'description': t.description,
            'category': t.category.name,
            'branch': t.branch.name if t.branch else '',
            'vendor': t.vendor_name,
            'debit': float(t.amount),
            'credit': 0.0,
            'receipt': 'Yes' if t.receipt else 'No',
            'status': 'Voided' if t.is_void else 'Posted',
        })
    for r in replenishments:
        entries.append({
            'date': r.date.isoformat(),
            'description': r.notes or 'Fund Replenishment',
            'category': f'Ref: {r.reference_number}' if r.reference_number else 'Replenishment',
            'branch': r.branch.name if r.branch else '',
            'vendor': '',
            'debit': 0.0,
            'credit': float(r.amount),
            'receipt': '',
            'status': 'Added',
        })

    entries.sort(key=lambda e: e['date'])

    total_debit = 0.0
    total_credit = 0.0
    for e in entries:
        ws.append([
            e['date'], e['description'], e['category'], e['branch'],
            e['vendor'], e['debit'], e['credit'], e['receipt'], e['status']
        ])
        if e['status'] != 'Voided':
            total_debit += e['debit']
        total_credit += e['credit']

    data_rows = len(entries) + 1
    for row_num in range(2, data_rows + 1):
        style_row(ws, row_num, len(headers), money_cols=[6, 7])

    # Footer row
    footer_row = data_rows + 1
    ws.cell(row=footer_row, column=1, value=f'Month Total — {month:02d}/{year}')
    ws.cell(row=footer_row, column=1).font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=footer_row, column=6, value=total_debit)
    ws.cell(row=footer_row, column=7, value=total_credit)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=footer_row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
        if col in [6, 7] and cell.value:
            cell.number_format = MONEY_FORMAT

    auto_width(ws, len(headers))
    return wb


def export_transactions(transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    headers = ['Date', 'Description', 'Category', 'Branch', 'Vendor',
               'Amount', 'Receipt', 'Status', 'Entered By']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    total = 0.0
    for t in transactions:
        ws.append([
            t.date.isoformat(), t.description, t.category.name,
            t.branch.name if t.branch else '', t.vendor_name,
            float(t.amount),
            'Yes' if t.receipt else 'No',
            'Voided' if t.is_void else 'Posted',
            t.entered_by.get_full_name() if t.entered_by else '',
        ])
        if not t.is_void:
            total += float(t.amount)

    data_rows = len(transactions) + 1
    for row_num in range(2, data_rows + 1):
        style_row(ws, row_num, len(headers), money_cols=[6])

    footer_row = data_rows + 1
    ws.cell(row=footer_row, column=1, value='Total')
    ws.cell(row=footer_row, column=1).font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=footer_row, column=6, value=total)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=footer_row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
        if col == 6 and cell.value:
            cell.number_format = MONEY_FORMAT

    auto_width(ws, len(headers))
    return wb


def export_replenishments(replenishments):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Replenishments'

    headers = ['Date', 'Amount', 'Reference #', 'Notes', 'Branch', 'Added By']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    total = 0.0
    for r in replenishments:
        ws.append([
            r.date.isoformat(), float(r.amount),
            r.reference_number or '', r.notes or '',
            r.branch.name if r.branch else '',
            r.added_by.get_full_name() if r.added_by else '',
        ])
        total += float(r.amount)

    data_rows = len(replenishments) + 1
    for row_num in range(2, data_rows + 1):
        style_row(ws, row_num, len(headers), money_cols=[2])

    footer_row = data_rows + 1
    ws.cell(row=footer_row, column=1, value='Total')
    ws.cell(row=footer_row, column=1).font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=footer_row, column=2, value=total)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=footer_row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
        if col == 2 and cell.value:
            cell.number_format = MONEY_FORMAT

    auto_width(ws, len(headers))
    return wb
